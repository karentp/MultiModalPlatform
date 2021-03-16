from django.shortcuts import render,redirect,get_object_or_404
from django.contrib.auth.models import User
from django.contrib.auth import authenticate,logout as logoutUser,login as loginUser
from django.contrib.auth.hashers import make_password
from django.http import Http404, HttpResponse
from django.core.files.base import ContentFile
from .models import *
from django.contrib import messages
from .filters import SegmentationFilter, CorpusFilter
from PIL import Image
import datetime
import xlwt
import os
from io import BytesIO, StringIO
import xlsxwriter
from .resources import SegmentationResource
from tablib import Dataset
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from uuid import uuid4
from .forms import CorpusForm

# Create your views here.

global segmentationfiltered

global corpusfiltered

def home(request):
    return render(request,"home.html",{"title":"Home","cssfile":"Home"})

def aboutus(request):
    return render(request,"about.html",{"title":"About us","cssfile":"contact"})


def segmentation(request):
    try:
        global segmentationfiltered
        segmentation=Segmentation.objects.all()
        total_segmentations = segmentation.count()
        
        myFilter = SegmentationFilter(request.GET, queryset=segmentation)
        segmentation = myFilter.qs
        segmentationfiltered = segmentation 
        total_segmentations = myFilter.qs.count()

        
        return render(request,"segmentation.html",{"title":"Segmentation","cssfile":"Segmentation",
                     "segmentations":segmentation, 'total_segmentations': total_segmentations, 
                     'myFilter': myFilter})
        
    except Exception as e:
            print(e)
            messages.error(request, "Something went to wrong",e)                
            return redirect('home')


def form(request):
    try:
        if request.user.is_authenticated:
            if request.method=='GET':
                return render(request,"addform.html",{"title":"Add Form","cssfile":"AddForm"})
            elif request.method=='POST':
                data =request.POST
                print(request.POST.get('code'))
                print(data)
                storeData = Segmentation.objects.create(color=data['color'],document_name=data['document_name'],code=data['code'],start=data['start'],end=data['end'],weight=data['weight'],image=request.FILES['image'],percentage=data['percentage'],condificados=data['condificados'],memos=data['memos'],created_by=request.user)
               
                print(storeData)
                print("IMAGEN ", storeData.image)
                areaSize=storeData.areaSize()
                print("areaSize",areaSize)
                storeData.area_size=areaSize
                storeData.save()
                messages.success(request, "Imagen guardada éxitosamente")                
                print("storeData",storeData.area_size)
                return redirect('segmentation')
        else:
            return redirect('segmentation')
    except Exception as e:
            print(e)
            messages.error(request, "Something went to wrong")                
            return redirect('segmentation')


def singleFormView(request,id):
    try:
        segmentation=Segmentation.objects.get(id=id)
        return render(request,"singleformview.html",{"title":"View Form","cssfile":"Segmentation","segmentation":segmentation})
    except Exception as e:
        print(e)
        return redirect('segmentation')
def login(request):
    try:
        if request.user.is_authenticated:
            return redirect('home')
        else:    
            if request.method=='GET':
                return render(request,"login.html",{"title":"Login","cssfile":"loginform"})
            elif request.method=='POST':
                email = request.POST['email']
                password = request.POST['password']        
                user = get_object_or_404(User,email=email)
                user = authenticate(username=user.username, password=password)
                loginUser(request,user)
                print(user)
                if user is not None:
                    print(user)
                    messages.success(request, "Inicio de sesión existoso")
                    return redirect('home')
                else:
                    messages.error(request, "Email or Password wrong")
                    return redirect('login')
        return redirect('home')

    except User.DoesNotExist:
        messages.error(request, "El usuario no existe.")
        print("User Not Found")
        return redirect('login')   
    except Exception as e:
        print(e)
        messages.error(request, "Something went to wrong try after sometime")
        return redirect('login')

def singup(request):
    try:
        if request.user.is_authenticated:
            return redirect('home')
        else:    
            if request.method=='GET':
                return render(request,"login.html",{"title":"Singup","cssfile":"Singup"})
            elif request.method=='POST':
                data =request.POST
                name = data['name']
                username = data['username']
                email = data['email1']
                password = data['password1']
                # data.password2=password
                user = User.objects.filter(email=email)
                if user:
                    messages.error(request,"La dirección de correo electrónico ya existe") 
                    return redirect('login')
                else:
                    user = User.objects.create(first_name= name,username=username,email=email,password=make_password(password))
                    
                    print(user)
                    messages.success(request, "Usuario registrado exitosamente")                
                    user.profile.institute = data['institute']
                    user.save()
                    return redirect('login')
    except Exception as e:
            print(e)
            messages.error(request, "Algo salió mal. Intente nuevamente.")                
            return redirect('login')


def logout(request):
    logoutUser(request)
    return redirect('home')


def export_excel(request):
    global segmentationfiltered
    output = BytesIO()
    
    workbook = xlsxwriter.Workbook(output, {'in_memory': True, 'remove_timezone': True})
    worksheet = workbook.add_worksheet('Datos')
    format= workbook.add_format({'bold': True, 'font_size': 16, 'align':'center'})

    workbook.formats[0].set_font_size(16)
    # Write the title for every column in bold
    worksheet.write('A1', 'Imagen', format)
    worksheet.write('B1', 'Nombre', format)
    worksheet.write('C1', 'Categoría', format)
    worksheet.write('D1', 'Fecha de creación',format)

    rows = segmentationfiltered.values_list(
        'image','document_name','code','created_at')

    # Start from the first cell. Rows and columns are zero indexed.
    row = 1
    col = 0
    cell_width = 45
    cell_height = 40
    worksheet.set_column('A:A', cell_width )
    worksheet.set_column('B:W', 25)
    worksheet.set_default_row(cell_height )
    # Iterate over the data and write it out row by row.
    format2= workbook.add_format({'bold': False, 'font_size': 14, 'align':'center'})


    for segment in rows:
        img = os.path.dirname(os.path.realpath('media'))+'/app/media/'+segment[0]
        imgobj = Image.open(img)
        width, height = imgobj.size
        print("SIZE", width, ' ', height)
        x_scale = 1
        y_scale = 1
        if(width > cell_width):
            x_scale = 1/(width/(cell_width*5))
        
        if(height > cell_height):
            y_scale = ((cell_height)/height)

        print("scale", x_scale, ' ', y_scale)
        worksheet.insert_image(row, col, img, {'x_scale': x_scale, 'y_scale': y_scale, 'align':'center', 'x_offset': 100, 'y_offset': 5})
        worksheet.write(row, col + 1, segment[col + 1],format2)
        worksheet.write(row, col + 2, segment[col + 2],format2)
        worksheet.write(row, col + 3, segment[col + 3],format2)
        row += 1

    workbook.close()
   
   

    output.seek(0)
    response = HttpResponse(output.read(),content_type ='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Segmentos' + \
        str(datetime.datetime.now())+'.xlsx'
    
    return response


def simple_upload(request):
    corpus_form = CorpusForm()
    if request.method == 'POST':
        corpus_form = CorpusForm(request.POST, request.FILES)
        if corpus_form.is_valid():
            corpus_form.save()
    context = {'corpus_form': corpus_form}
    return render(request,'upload.html', context)
            
def corpus_upload(request,id):
    corpus=Corpus.objects.get(id=id)
    corpus_form = CorpusForm(instance = corpus)
    if request.method == 'POST':
        corpus_form = CorpusForm(request.POST, request.FILES, instance = corpus)
        if corpus_form.is_valid():
            corpus_form.save()
        
        context = {'corpus_form': corpus_form}
        segmentation=Segmentation.objects.all()

        print("EJEMPLO ", segmentation[1].image)
        
        dataset = Dataset()
        #new_segmentations = request.FILES['myfile']
        new_segmentations = corpus.corpus_document
        imported_data = dataset.load(new_segmentations.read(), format='xlsx')
        rowNumber = 2
        print('PATH', request.FILES['myfile'])
        wb = load_workbook(request.FILES['myfile'])
        sheet = wb['Datos']

        image_loader = SheetImageLoader(sheet)
        rowNumber = 2
        for data in imported_data:
            
            cell = str((sheet.cell(row=rowNumber, column=1)).coordinate)
            value = Segmentation.objects.create(image=data[0],document_name=data[1],code=data[2],created_by=request.user)
            value.corpus = corpus
            value.save()
            if image_loader.image_in(cell):
                im = image_loader.get(cell)
                print(sheet.cell(row=rowNumber, column=1).value)
                print(im)
                if(im.mode != "RGB"):
                    im = im.convert("RGB")
                #image.save(data[1]+'.jpg')
                #value.image.save(data[1]+'.jpg',image, save = False) 
                #im = Image.open(image)
                #im.thumbnail((220, 130), Image.ANTIALIAS)
                print("before thumb")
                thumb_io = BytesIO()
                print("THUMB ", thumb_io)
                print("FORMAT", im.format)
                filename = '{}.{}'.format(uuid4().hex,'jpg')
                print("FILENAME",filename)
                im.save('app/media/upload_images/'+filename)
                print("after thumb")
                value.image='upload_images/'+filename
                #value.image.save(im.filename, ContentFile(thumb_io.getvalue()), save=False)
                value.save()
                print("DATOS",data[0]," ",data[1]," ",data[2]," ",data[3]," ")
                print("IMG VALUE ", value.image)
            value.save()
            print("VALUE", value)
            rowNumber +=1

    return render(request,'corpus_upload.html', context)
    

def corpus_listing(request):
    try:
        global corpusfiltered
        corpus=Corpus.objects.all()
        total_corpus = corpus.count()
        print("antes filtro")
        myFilter = CorpusFilter(request.GET, queryset=corpus)
        print("desp filtro")
        corpus = myFilter.qs
        corpusfiltered = corpus
        print("global")
        total_corpus = myFilter.qs.count()

        return render(request,"corpus.html",{"title":"Corpus","cssfile":"Segmentation",
                     "corpus":corpus, 'total_corpus': total_corpus, 
                     'myFilter': myFilter})
        
    except Exception as e:
            print(e)
            messages.error(request, "Hubo un error, intente de nuevo.",e)                
            return redirect('home')

def corpus_for_approval(request):
    try:
        global corpusfiltered
        corpus=Corpus.objects.all()
        total_corpus = corpus.count()
        myFilter = CorpusFilter(request.GET, queryset=corpus)
        corpus = myFilter.qs
        corpusfiltered = corpus
        total_corpus = myFilter.qs.count()

        return render(request,"corpus_for_approval.html",{"title":"Corpus","cssfile":"Segmentation",
                     "corpus":corpus, 'total_corpus': total_corpus, 
                     'myFilter': myFilter})
        
    except Exception as e:
        messages.error(request, "Hubo un error, intente de nuevo.",e)                
        return redirect('home')

def singleCorpusView(request, id):
    try:
        corpus=Corpus.objects.get(id=id)
        corpus_form = CorpusForm(instance = corpus)
        for fieldname in corpus_form.fields:
            corpus_form.fields[fieldname].disable = True
        context = {'corpus_form': corpus_form, 'corpus':corpus}
        return render(request,"single_corpus_view.html",context)
    except Exception as e:
        return redirect('corpus_listing')